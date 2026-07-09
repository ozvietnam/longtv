#!/usr/bin/env python3
"""Generate All-in-One Year-1 CAPEX/OPEX Excel — HP + TN branches."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "03-departments/05-van-hanh-tc/financial-plan-all-in-one-year1.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="AAAAAA")


def style_header(ws, row, cols, text):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        if c == 1:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = HEADER_FILL
        else:
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=1, value=text)


def write_table(ws, start_row, headers, rows, hp_col=2, tn_col=3, total_col=4):
    cols = len(headers)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = SUBHEADER_FILL
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    r = start_row + 1
    hp_sum = tn_sum = 0
    for label, hp, tn, note in rows:
        ws.cell(row=r, column=1, value=label).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        ws.cell(row=r, column=hp_col, value=hp).number_format = "#,##0"
        ws.cell(row=r, column=tn_col, value=tn).number_format = "#,##0"
        ws.cell(row=r, column=total_col, value=f"=SUM({get_column_letter(hp_col)}{r}:{get_column_letter(tn_col)}{r})").number_format = "#,##0"
        if len(headers) > 4:
            ws.cell(row=r, column=5, value=note or "")
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        hp_sum += hp
        tn_sum += tn
        r += 1

    # Total row
    for c in range(1, cols + 1):
        ws.cell(row=r, column=c).font = Font(bold=True)
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ws.cell(row=r, column=1, value="TỔNG CỘNG")
    ws.cell(row=r, column=hp_col, value=f"=SUM({get_column_letter(hp_col)}{start_row+1}:{get_column_letter(hp_col)}{r-1})").number_format = "#,##0"
    ws.cell(row=r, column=tn_col, value=f"=SUM({get_column_letter(tn_col)}{start_row+1}:{get_column_letter(tn_col)}{r-1})").number_format = "#,##0"
    ws.cell(row=r, column=total_col, value=f"=SUM({get_column_letter(total_col)}{start_row+1}:{get_column_letter(total_col)}{r-1})").number_format = "#,##0"
    return r, hp_sum, tn_sum


CAPEX_ROWS = [
    ("Giấy phép & pháp lý thương hiệu (luật liên kết, mã đại lý HQ, bưu chính)", 185_000_000, 140_000_000, "MOU công ty luật + đại lý ECUS"),
    ("Hạ tầng văn phòng (cọc 3–6 tháng + nội thất tiếp khách chuẩn TQ)", 385_000_000, 295_000_000, "HP: Lê Hồng Phong / Đình Vũ · TN: Yên Bình"),
    ("CNTT & phần mềm (WMS, CRM, ECUS5-VNACCS bản quyền/năm 1)", 275_000_000, 180_000_000, "Zoho/HubSpot + đối tác WMS"),
]

OPEX_ROWS = [
    ("Quỹ lương — đội pháp lý (IRC, ERC, ĐTM, PCCC)", 960_000_000, 720_000_000, "HP 4 FTE · TN 3 FTE"),
    ("Quỹ lương — logistics hiện trường & chứng từ (HQ, kho)", 1_080_000_000, 648_000_000, "Chứng chỉ TCHQ · song ngữ"),
    ("Quỹ lương — BD/PM (HSK 5–6, Guanxi)", 360_000_000, 312_000_000, ""),
    ("Thuê kho bãi & bốc xếp", 336_000_000, 280_000_000, "HP consolidation cảng · TN KCN Yên Bình"),
    ("Marketing TQ (Baidu, WeChat, hội thảo Việt–Trung)", 80_000_000, 60_000_000, ""),
    ("Thuê VP (phần OPEX, sau cọc) + điện nước", 36_000_000, 30_000_000, ""),
    ("Tuyển dụng / HR cung ứng LĐ (TN)", 0, 150_000_000, "Samsung hub competition"),
    ("SaaS, đi lại, admin, dự phòng phát sinh", 24_000_000, 114_000_000, ""),
]

# Verify sums
assert sum(x[1] for x in CAPEX_ROWS) == 845_000_000
assert sum(x[2] for x in CAPEX_ROWS) == 615_000_000
assert sum(x[1] for x in OPEX_ROWS) == 2_876_000_000
assert sum(x[2] for x in OPEX_ROWS) == 2_314_000_000


def build_capex_sheet(wb):
    ws = wb.create_sheet("CAPEX")
    ws.column_dimensions["A"].width = 52
    for col in "BCD":
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["E"].width = 36
    ws["A1"] = "CHI PHÍ ĐẦU TƯ BAN ĐẦU (CAPEX) — NĂM 1"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    write_table(ws, 3, ["Hạng mục", "Hải Phòng (VND)", "Thái Nguyên (VND)", "Tổng (VND)", "Ghi chú"], CAPEX_ROWS)
    return ws


def build_opex_sheet(wb):
    ws = wb.create_sheet("OPEX")
    ws.column_dimensions["A"].width = 52
    for col in "BCD":
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["E"].width = 36
    ws["A1"] = "CHI PHÍ VẬN HÀNH NĂM ĐẦU (OPEX) — NĂM 1"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    write_table(ws, 3, ["Hạng mục", "Hải Phòng (VND)", "Thái Nguyên (VND)", "Tổng (VND)", "Ghi chú"], OPEX_ROWS)
    return ws


def build_overview(wb):
    ws = wb.active
    ws.title = "Tổng Quan Dự Án"
    ws.column_dimensions["A"].width = 42
    for col in "BCDE":
        ws.column_dimensions[col].width = 20

    ws["A1"] = "LONGTV — KẾ HOẠCH TÀI CHÍNH NĂM 1 (ALL-IN-ONE)"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:E1")
    ws["A2"] = "Quy mô tầm trung · 2 chi nhánh HP + TN · Giả định desk 2026-07-09"
    ws.merge_cells("A2:E2")

    headers = ["Hạng mục chi phí", "Hải Phòng (VND)", "Thái Nguyên (VND)", "Tổng hệ thống (VND)", "Tỷ trọng"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    rows = [
        ("1. CAPEX — Đầu tư ban đầu", "=CAPEX!B7", "=CAPEX!C7", "=CAPEX!D7"),
        ("2. OPEX — Vận hành năm đầu", "=OPEX!B12", "=OPEX!C12", "=OPEX!D12"),
        ("TỔNG NHU CẦU VỐN NĂM 1", "=B6+B7", "=C6+C7", "=D6+D7"),
    ]
    for i, (label, hp, tn, tot) in enumerate(rows, 6):
        ws.cell(row=i, column=1, value=label).font = Font(bold=(i == 8))
        ws.cell(row=i, column=2, value=hp).number_format = "#,##0"
        ws.cell(row=i, column=3, value=tn).number_format = "#,##0"
        ws.cell(row=i, column=4, value=tot).number_format = "#,##0"
        if i < 8:
            ws.cell(row=i, column=5, value=f"=D{i}/$D$8").number_format = "0.0%"
        for c in range(1, 6):
            ws.cell(row=i, column=c).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        if i == 8:
            for c in range(1, 5):
                ws.cell(row=i, column=c).fill = TOTAL_FILL

    ws["A10"] = "So sánh kịch bản"
    ws["A10"].font = Font(bold=True)
    compare = [
        ("Kịch bản bootstrap (QĐ #003)", 2_000_000_000, "Virtual office · 8–10 FTE · MOU đối tác"),
        ("Kịch bản tầm trung (file này)", "=D8", "2 VP + kho + ~16–18 FTE"),
        ("Chênh lệch cần bổ sung", "=D8-2000000000", "Gọi vốn vòng 2 hoặc giai đoạn 2–3"),
    ]
    for j, (k, v, note) in enumerate(compare, 11):
        ws.cell(row=j, column=1, value=k)
        cell = ws.cell(row=j, column=2, value=v)
        if isinstance(v, str) and v.startswith("="):
            cell.number_format = "#,##0"
        else:
            cell.number_format = "#,##0"
        ws.cell(row=j, column=3, value=note)
        ws.merge_cells(start_row=j, start_column=3, end_row=j, end_column=5)

    ws["A15"] = "Hướng dẫn"
    ws["A15"].font = Font(bold=True)
    notes = [
        "• Tab CAPEX / OPEX: sửa ô số liệu — Tổng tự cập nhật bằng SUM.",
        "• Tab Tổng Quan: đồng bộ qua tham chiếu =CAPEX! / =OPEX!.",
        "• Markdown: /docs/03-departments/05-van-hanh-tc/capex-opex-year1-medium-scale.md",
    ]
    for k, line in enumerate(notes, 16):
        ws.cell(row=k, column=1, value=line)
        ws.merge_cells(start_row=k, start_column=1, end_row=k, end_column=5)


def main():
    wb = Workbook()
    build_capex_sheet(wb)
    build_opex_sheet(wb)
    build_overview(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
